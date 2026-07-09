from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "WBS任务清单"

FONT = "Microsoft YaHei"
title_font = Font(name=FONT, bold=True, size=16, color="FFFFFF")
head_font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
cell_font = Font(name=FONT, size=10, color="000000")
note_font = Font(name=FONT, bold=True, size=10, color="000000")

title_fill = PatternFill("solid", fgColor="2F5496")
head_fill = PatternFill("solid", fgColor="4472C4")
note_fill = PatternFill("solid", fgColor="FFF2CC")

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

module_fills = {
    "基础设施": "E2EFDA", "数据库": "DEEBF7", "安全": "FCE4D6",
    "图书模块": "FFF2CC", "读者模块": "EDEDED", "借阅模块": "DDEBF7", "测试": "F8CBAD",
}

headers = ["任务ID", "所属REQ", "任务名称", "估点", "依赖", "模块", "状态"]
rows = [
    ["T-01", "REQ-03", "初始化 Spring Boot 2.7.18 项目 + PostgreSQL 配置（application.yml）", 2, "—", "基础设施", "待开始"],
    ["T-02", "REQ-03", "执行 DDL 建表（books/readers/borrow_records/audit_logs）+ 索引", 3, "T-01", "数据库", "待开始"],
    ["T-03", "REQ-00", "实现 JWT 认证登录接口（POST /auth/login）：bcrypt 密码校验、凭证错误不暴露原因、连续5次失败锁定账户15分钟、token 携带 readerId+role", 4, "T-02", "安全", "待开始"],
    ["T-14", "NFR-10", "全局异常处理 + 统一响应体（Result<T>）", 2, "T-01", "基础设施", "待开始"],
    ["T-15", "NFR-05", "JWT 拦截器 + 角色权限校验（ROLE_READER / CIRCULATION / ADMIN）", 3, "T-03", "安全", "待开始"],
    ["T-06", "REQ-10", "管理员上架图书接口（POST /admin/books）", 2, "T-02", "图书模块", "待开始"],
    ["T-07", "REQ-11", "管理员下架图书接口（DELETE /admin/books/{isbn}，逻辑删除）", 2, "T-02", "图书模块", "待开始"],
    ["T-04", "REQ-01", "图书检索接口（分页 + 模糊搜索：书名/作者/ISBN）", 5, "T-02", "图书模块", "待开始"],
    ["T-05", "REQ-02", "图书详情接口（GET /books/{isbn}）", 2, "T-02", "图书模块", "待开始"],
    ["T-08", "REQ-06", "馆员注册读者接口（POST /readers，默认密码=手机号后6位）", 3, "T-02", "读者模块", "待开始"],
    ["T-09", "REQ-07", "修改读者手机号接口（PUT /readers/{readerId}/phone）", 2, "T-08", "读者模块", "待开始"],
    ["T-10", "REQ-08", "管理员重置读者密码接口（PUT /admin/readers/{readerId}/reset-password）", 2, "T-08", "读者模块", "待开始"],
    ["T-11", "REQ-03", "借书接口（POST /borrow-records，含超期检查/上限校验/库存扣减/乐观锁防并发）", 5, "T-02, T-04, T-08", "借阅模块", "待开始"],
    ["T-12", "REQ-04", "还书接口（POST /borrow-records/{recordId}/return，含库存释放/超期计算/自动解冻）", 3, "T-11", "借阅模块", "待开始"],
    ["T-13", "REQ-05", "读者在借清单接口（GET /readers/{readerId}/borrows）", 2, "T-08, T-11", "借阅模块", "待开始"],
    ["T-16", "—", "核心流程集成测试（各角色登录→张三注册→管理员上架图书→读者检索→李四借书→查看在借→王五还书→张三超期被拒→还书后解冻）", 3, "T-12", "测试", "待开始"],
]

ncols = len(headers)
# Title row
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
c = ws.cell(row=1, column=1, value="EvoLib 图书馆管理 MVP 系统 —— WBS 任务清单")
c.font = title_font; c.fill = title_fill; c.alignment = center
ws.row_dimensions[1].height = 30

# Subtitle
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
c = ws.cell(row=2, column=1, value="依据 SRS_v1.0（REQ-00 ~ REQ-11）| 16 个任务 | 总估点 45 点 | 计划版本 V1.0 | 2026-07-09")
c.font = Font(name=FONT, size=10, italic=True, color="595959")
c.alignment = center
ws.row_dimensions[2].height = 20

# Header row
hr = 3
for j, h in enumerate(headers, 1):
    c = ws.cell(row=hr, column=j, value=h)
    c.font = head_font; c.fill = head_fill; c.alignment = center; c.border = border
ws.row_dimensions[hr].height = 22

# Data rows
r = hr + 1
for row in rows:
    for j, v in enumerate(row, 1):
        c = ws.cell(row=r, column=j, value=v)
        c.font = cell_font; c.border = border
        c.alignment = left if j == 3 else center
    mf = module_fills.get(row[5])
    if mf:
        ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor=mf)
    ws.row_dimensions[r].height = 40
    r += 1

# Total row
ws.cell(row=r, column=1, value="总估点")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
tc = ws.cell(row=r, column=1); tc.font = note_font; tc.fill = note_fill; tc.alignment = center; tc.border = border
sc = ws.cell(row=r, column=4, value=45)
sc.font = note_font; sc.fill = note_fill; sc.alignment = center; sc.border = border
ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
nc = ws.cell(row=r, column=5, value="单人 1 周冲刺，每日有效产出约 6-8 点，严格锁定范围")
nc.font = note_font; nc.fill = note_fill; nc.alignment = center; nc.border = border
for col in range(1, ncols+1):
    ws.cell(row=r, column=col).border = border
ws.row_dimensions[r].height = 24

widths = [10, 10, 58, 8, 16, 12, 10]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w

ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A{hr}:G{r-1}"

wb.save(r"E:\ai_fq\evolib\docs\WBS_v1.0.xlsx")
print("saved")
